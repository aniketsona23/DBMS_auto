import os
import json
import re
import zipfile
from pathlib import Path
from typing import Dict
from typing import Any
from openpyxl import Workbook, load_workbook


from shared.constants import FieldNames, KEY_PATH
from shared.logger import get_logger
from shared.encryption import decrypt_data, get_or_create_key
from shared.models import QuestionResult, format_question_for_excel

logger = get_logger(__name__)


def extract_student_id(filename: str) -> str:
    """
    Extract student ID from filename pattern: YYYY[A-Z][A-Z or 0-9][A-Z][A-Z or 0-9][0-9][0-9][0-9][0-9]G
    Example: 2024AB12C3456G -> 2024AB12C3456G
    """
    # Pattern: 4 digits + [A-Z] + [A-Z0-9] + [A-Z] + [A-Z0-9] + 4 digits + G
    pattern = r"(\d{4}[A-Z][A-Z0-9][A-Z][A-Z0-9]\d{4}G)"
    match = re.search(pattern, filename, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return ""


def find_and_decrypt_results(inner_zip_path: Path, key: bytes) -> dict:
    """Find and decrypt results.json from inner ZIP file."""
    # Prefer files that look like results reports (e.g., '*_results.json.enc')
    with zipfile.ZipFile(inner_zip_path, "r") as inner_zip:
        namelist = inner_zip.namelist()
        # First try to find a results file explicitly
        results_files = [
            f for f in namelist if re.search(r"results\.json\.enc$", f, re.I)
        ]
        candidates = (
            results_files
            if results_files
            else [f for f in namelist if f.endswith(".json.enc")]
        )

        for file_info in candidates:
            data = inner_zip.read(file_info)
            try:
                # Try decryption first
                try:
                    decrypted = decrypt_data(data, key)
                    decoded_json = json.loads(decrypted.decode("utf-8"))
                except Exception:
                    # If decryption fails, try plain JSON
                    decoded_json = json.loads(data.decode("utf-8"))
                return decoded_json
            except Exception as e:
                logger.warning(f"Failed to decode {file_info}: {e}")
    return {}


def process_zip_file(zip_path: Path, key: bytes, excel_path: str) -> bool:
    """Process a single outer ZIP file and extract results."""
    filename = zip_path.name
    student_id_from_zip = extract_student_id(filename)

    if not student_id_from_zip:
        logger.warning(f"Could not extract student ID from: {filename}")
        return False

    logger.info(f"Processing {filename} -> Student ID: {student_id_from_zip}")

    try:
        with zipfile.ZipFile(zip_path, "r") as outer_zip:
            # Look for inner ZIP file
            inner_zip_files = [f for f in outer_zip.namelist() if f.endswith(".zip")]

            if not inner_zip_files:
                logger.warning(f"No inner ZIP found in {filename}")
                return False

            # Extract first inner ZIP to temp location
            inner_zip_name = inner_zip_files[0]
            inner_zip_data = outer_zip.read(inner_zip_name)
            temp_inner = Path(f"temp_{student_id_from_zip}.zip")
            temp_inner.write_bytes(inner_zip_data)

            try:
                # Find and decrypt results.json
                results = find_and_decrypt_results(temp_inner, key)

                if not results:
                    logger.warning(f"No results found in {filename}")
                    return False

                uploaded_id = results.get(FieldNames.STUDENT_ID, student_id_from_zip)
                timestamp = results.get(FieldNames.TIMESTAMP, "")
                total_score = results.get(FieldNames.TOTAL_SCORE, 0)
                max_score = results.get(FieldNames.MAX_SCORE, 0)

                # Build question feedback dictionary using standardized models
                questions_data: Dict[str, QuestionResult] = results.get(
                    FieldNames.QUESTIONS, {}
                )
                # Only include keys that look like question keys (q1, q2, ...)
                questions = {
                    test_key: format_question_for_excel(test_info)
                    for test_key, test_info in questions_data.items()
                    if isinstance(test_key, str) and test_key.startswith("q")
                }

                append_to_excel(
                    excel_path=excel_path,
                    student_id_zip=student_id_from_zip,
                    uploaded_id=uploaded_id,
                    timestamp=timestamp,
                    total_score=total_score,
                    questions=questions,
                )

                logger.info(
                    f"[SUCCESS] Processed {student_id_from_zip}: {total_score}/{max_score} at {timestamp}"
                )
                return True

            finally:
                # Cleanup temp file
                if temp_inner.exists():
                    temp_inner.unlink()

    except Exception as e:
        logger.error(f"[ERROR] Failed to process {filename}: {e}")
        return False


def append_to_excel(
    excel_path: str,
    student_id_zip: str,
    uploaded_id: str,
    timestamp: str,
    total_score: float,
    questions: dict,
):
    """Append results to Excel file."""
    # Prepare headers
    base_cols = [FieldNames.STUDENT_ID, "uploaded_id", FieldNames.TIMESTAMP]

    # Sort question keys numeric q-keys (q1, q2...) by numeric order
    def _question_sort_key(k: str):
        if k.startswith("q") and k[1:].isdigit():
            return (0, int(k[1:]))
        return (1, k)

    question_keys = sorted(questions.keys(), key=_question_sort_key)
    question_cols = [qk for qk in question_keys]

    all_cols = base_cols + question_cols + [FieldNames.TOTAL_SCORE]

    # Prepare row data
    # Ensure all row values are strings to satisfy Excel writer typing
    new_row = [str(student_id_zip), str(uploaded_id), str(timestamp)]
    for qk in question_keys:
        new_row.append(str(questions.get(qk, "")))
    new_row.append(str(total_score))

    # Load or create workbook
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        ws_any: Any = ws

        # Update headers if needed
        existing = [cell.value for cell in ws_any[1]]
        for idx, col in enumerate(all_cols):
            if idx >= len(existing) or existing[idx] != col:
                ws_any.cell(row=1, column=idx + 1).value = col
    else:
        wb = Workbook()
        ws = wb.active
        ws_any: Any = ws
        for idx, col in enumerate(all_cols):
            ws_any.cell(row=1, column=idx + 1).value = col

    # Append row
    ws_any.append(new_row)
    wb.save(excel_path)
    logger.info(
        f"Appended to {excel_path}: {student_id_zip} -> total_score: {total_score}"
    )


def main():
    """Main function to process all ZIP files in current directory."""
    excel_path = "grades.xlsx"

    key = get_or_create_key(KEY_PATH)

    # Find all ZIP files in current directory
    zip_files = sorted(Path(".").glob("*.zip"))

    if not zip_files:
        logger.info("No ZIP files found in current directory")
        return

    logger.info(f"Found {len(zip_files)} ZIP file(s) to process")
    logger.info(f"Output will be saved to: {excel_path}")
    logger.info("-" * 60)

    success_count = 0
    for zip_path in zip_files:
        if process_zip_file(zip_path, key, excel_path):
            success_count += 1
        logger.info("")  # Empty line

    logger.info("-" * 60)
    logger.info(f"Successfully processed {success_count}/{len(zip_files)} ZIP files")
    logger.info(f"Results saved to: {excel_path}")


if __name__ == "__main__":
    main()
