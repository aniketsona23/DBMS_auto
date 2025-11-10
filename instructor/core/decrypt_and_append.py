import os
import sys
import json
from openpyxl import Workbook, load_workbook
from cryptography.fernet import Fernet


def get_key():
    key = os.environ.get("FERNET_KEY")
    if not key:
        key = b"nVPZZCjg6EFcWrch2Ivk13WWXNv7uWZGU5C5Vc2ADrw="
    return key.encode() if isinstance(key, str) else key


def main():
    report_path = input(
        "Enter path to encrypted results file (default: results.json.enc): "
    ).strip()
    if not report_path:
        report_path = "results.json.enc"
    excel_path = "grades.xlsx"
    if not os.path.exists(report_path):
        print(f"[ERROR] File not found: {report_path}")
        sys.exit(1)
    key = get_key()
    with open(report_path, "rb") as f:
        enc = f.read()
    fernet = Fernet(key)
    try:
        data = json.loads(fernet.decrypt(enc).decode("utf-8"))
    except Exception as e:
        print(f"[ERROR] Could not decrypt or parse: {e}")
        sys.exit(1)
    # New minimal schema: { student_id, timestamp, total_score, questions: { q1: "Pass" | reason } }
    student_id = data.get("student_id", "")
    timestamp = data.get("timestamp", "")
    total_score = data.get("total_score", "")
    questions = data.get("questions", {})

    # Prepare headers: student_id, timestamp, q1..qN, total_score
    base_cols = ["student_id", "timestamp"]
    question_keys = sorted(
        questions.keys(),
        key=lambda k: int(k[1:]) if k.startswith("q") and k[1:].isdigit() else k,
    )
    all_cols = base_cols + question_keys + ["total_score"]

    new_row = [student_id, timestamp]
    new_row += [questions.get(k, "") for k in question_keys]
    new_row += [total_score]
    # Load or create workbook
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        # If headers don't match, extend as needed (dynamic columns)
        existing = [cell.value for cell in ws[1]]
        for idx, col in enumerate(all_cols):
            if idx >= len(existing) or existing[idx] != col:
                ws.cell(row=1, column=idx + 1).value = col
    else:
        wb = Workbook()
        ws = wb.active
        for idx, col in enumerate(all_cols):
            ws.cell(row=1, column=idx + 1).value = col
    # Append row
    ws.append(new_row)
    wb.save(excel_path)
    print(f"[INFO] Data appended for student id {student_id}; file: {excel_path}")
    print(f"  Total Score: {total_score} at {timestamp}")
    print("  Question outcomes:")
    for k in question_keys:
        print(f"    {k}: {questions.get(k, '')}")


if __name__ == "__main__":
    main()
