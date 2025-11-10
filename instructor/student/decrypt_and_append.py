import os
import sys
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from cryptography.fernet import Fernet


def get_key():
    key = os.environ.get("FERNET_KEY")
    if not key:
        key = input("Enter Fernet key (base64, as in ENCRYPTION_KEY): ").strip()
    return key.encode() if isinstance(key, str) else key


def main():
    report_path = input(
        "Enter path to encrypted report.json.enc (default: report.json.enc): "
    ).strip()
    if not report_path:
        report_path = "report.json.enc"
    excel_path = "grades.xlsx"
    if not os.path.exists(report_path):
        print(f"[ERROR] File not found: {report_path}")
        sys.exit(1)
    key = get_key()
    with open(report_path, "rb") as f:
        enc = f.read()
    fernet = Fernet(key)
    try:
        data = json.loads(fernet.decrypt(enc).decode("utf-8"))
    except Exception as e:
        print(f"[ERROR] Could not decrypt or parse: {e}")
        sys.exit(1)
    # Extract basic info
    student_id = data.get("student_id", "")
    timestamp = data.get("timestamp", "")
    total_score = data.get("total_score", "")
    max_score = data.get("max_score", "")
    percentage = data.get("percentage", "")
    test_results = data.get("test_results", [])

    # Header columns - now includes status, score, max_score, and feedback for each test
    cols = ["student_id", "timestamp", "total_score", "max_score", "percentage"]
    test_ids = [tr.get("test") for tr in test_results]

    # Create columns for status, score, max_score, and feedback for each question
    status_cols = [f"{tid}_status" for tid in test_ids]
    score_cols = [f"{tid}_score" for tid in test_ids]
    max_score_cols = [f"{tid}_max_score" for tid in test_ids]
    feedback_cols = [f"{tid}_feedback" for tid in test_ids]

    all_cols = cols + status_cols + score_cols + max_score_cols + feedback_cols

    # Build the data row
    new_row = [student_id, timestamp, total_score, max_score, percentage]

    # Add status for each test
    new_row += [tr.get("status") for tr in test_results]

    # Add score for each test
    new_row += [tr.get("score") for tr in test_results]

    # Add max_score for each test
    new_row += [tr.get("max_score") for tr in test_results]

    # Add feedback for each test (message + additional details)
    for tr in test_results:
        feedback = tr.get("message", "")

        # Add constraint violation details
        if "constraint" in feedback.lower():
            feedback = f"CONSTRAINT VIOLATION: {feedback}"

        # Add failure details for function tests
        if "failures" in tr:
            failures = tr.get("failures", [])
            feedback += f" | Failures: {'; '.join(failures)}"

        # Add mismatch details
        if "expected" in tr and "actual" in tr and tr.get("status") == "FAIL":
            expected = tr.get("expected", [])
            actual = tr.get("actual", [])
            feedback += f" | Expected rows: {len(expected)}, Got rows: {len(actual)}"

        new_row.append(feedback)
    # Load or create workbook
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        # If headers don't match, extend as needed (dynamic columns)
        existing = [cell.value for cell in ws[1]]
        for idx, col in enumerate(all_cols):
            if idx >= len(existing) or existing[idx] != col:
                ws.cell(row=1, column=idx + 1).value = col
    else:
        wb = Workbook()
        ws = wb.active
        for idx, col in enumerate(all_cols):
            ws.cell(row=1, column=idx + 1).value = col
    # Append row
    ws.append(new_row)
    wb.save(excel_path)
    print(f"[INFO] Data appended for student id {student_id}; file: {excel_path}")
    print(f"  Score: {total_score} / {max_score} = {percentage}% at {timestamp}")
    print("  Individual test results:")
    for tr in test_results:
        status = tr.get("status")
        score = tr.get("score")
        max_s = tr.get("max_score")
        msg = tr.get("message", "")
        test_id = tr.get("test", "unknown")

        # Color coding for terminal output
        if status == "PASS":
            status_icon = "✓"
        elif status == "FAIL":
            status_icon = "✗"
        elif status == "ERROR":
            status_icon = "⚠"
        else:
            status_icon = "?"

        print(f"    {status_icon} {test_id}: {status} ({score}/{max_s}) - {msg}")

        # Print additional details for failures
        if "failures" in tr:
            for failure in tr["failures"]:
                print(f"        - {failure}")

        if status == "FAIL" and "expected" in tr and "actual" in tr:
            expected_count = len(tr.get("expected", []))
            actual_count = len(tr.get("actual", []))
            print(f"        Expected {expected_count} rows, got {actual_count} rows")


if __name__ == "__main__":
    main()
